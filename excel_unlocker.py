#!/usr/bin/env python3
"""
Excel Sheet Unlocker - Cross Platform (Windows/macOS/Linux)
Removes sheet protection from Excel files without needing the password
Supports Windows, macOS, and Linux operating systems
"""

import openpyxl
import os
import sys
import platform
import subprocess
from pathlib import Path

def get_system_info():
    """Get system information for cross-platform compatibility"""
    system = platform.system().lower()
    return {
        'os': system,
        'is_windows': system == 'windows',
        'is_macos': system == 'darwin',
        'is_linux': system == 'linux',
        'python_cmd': 'python' if system == 'windows' else 'python3'
    }

def install_dependencies():
    """Install required packages cross-platform"""
    sys_info = get_system_info()
    python_cmd = sys_info['python_cmd']
    
    try:
        import openpyxl
        return True
    except ImportError:
        print("📦 Installing required package: openpyxl...")
        
        try:
            # Try pip install
            if sys_info['is_windows']:
                subprocess.check_call([python_cmd, '-m', 'pip', 'install', 'openpyxl'], 
                                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            else:
                subprocess.check_call([python_cmd, '-m', 'pip', 'install', 'openpyxl'], 
                                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except subprocess.CalledProcessError:
            try:
                # Fallback to direct pip
                subprocess.check_call(['pip', 'install', 'openpyxl'], 
                                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            except subprocess.CalledProcessError:
                print("❌ Failed to install openpyxl. Please install manually:")
                print(f"   {python_cmd} -m pip install openpyxl")
                return False
        
        print("✅ Installation complete!")
        return True

def normalize_path(path_str):
    """Normalize file paths for cross-platform compatibility"""
    if not path_str:
        return path_str
    
    # Remove quotes that might be added by drag-and-drop
    path_str = path_str.strip().strip('"').strip("'")
    
    # Convert to Path object for cross-platform handling
    path_obj = Path(path_str)
    
    # Resolve to absolute path
    try:
        return str(path_obj.resolve())
    except:
        return str(path_obj.expanduser().resolve())
    """
    Remove protection from Excel sheets without password
    
    Args:
        file_path (str): Path to the protected Excel file
        output_path (str, optional): Path for unlocked file. If None, adds '_unlocked' suffix
    
    Returns:
        bool: True if successful, False otherwise
    """
def unlock_excel_sheets(file_path, output_path=None):
    """
    Remove protection from Excel sheets without password
    
    Args:
        file_path (str): Path to the protected Excel file
        output_path (str, optional): Path for unlocked file. If None, adds '_unlocked' suffix
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Normalize paths for cross-platform compatibility
        file_path = normalize_path(file_path)
        
        # Check if file exists
        if not os.path.exists(file_path):
            print(f"❌ Error: File '{file_path}' not found.")
            return False
        
        print(f"📂 Loading Excel file: {os.path.basename(file_path)}")
        
        # Load workbook (this bypasses sheet-level password protection)
        workbook = openpyxl.load_workbook(file_path)
        
        # Process all sheets
        unlocked_sheets = []
        total_sheets = len(workbook.sheetnames)
        
        for sheet_name in workbook.sheetnames:
            print(f"🔍 Processing sheet: '{sheet_name}'")
            worksheet = workbook[sheet_name]
            
            try:
                # Method 1: Direct protection removal
                from openpyxl.worksheet.protection import SheetProtection
                
                # Check current protection status
                was_protected = False
                try:
                    was_protected = bool(worksheet.protection.sheet)
                except:
                    was_protected = True  # Assume protected if we can't determine
                
                # Always create a fresh, unprotected SheetProtection object
                new_protection = SheetProtection(
                    sheet=False,
                    password=None,
                    selectLockedCells=True,
                    selectUnlockedCells=True,
                    formatCells=True,
                    formatColumns=True,
                    formatRows=True,
                    insertColumns=True,
                    insertRows=True,
                    insertHyperlinks=True,
                    deleteColumns=True,
                    deleteRows=True,
                    sort=True,
                    autoFilter=True,
                    pivotTables=True,
                    objects=True,
                    scenarios=True
                )
                
                # Replace the protection object entirely
                worksheet.protection = new_protection
                
                if was_protected:
                    print(f"🔓 Successfully unlocked sheet: '{sheet_name}'")
                    unlocked_sheets.append(sheet_name)
                else:
                    print(f"✅ Sheet '{sheet_name}' was already unlocked")
                    
            except Exception as e:
                print(f"⚠️  Method 1 failed for '{sheet_name}': {e}")
                
                # Method 2: Brute force approach
                try:
                    # Try to unlock by directly modifying the worksheet's XML
                    worksheet._protection = None
                    
                    # Create minimal protection object
                    from openpyxl.worksheet.protection import SheetProtection
                    worksheet.protection = SheetProtection()
                    worksheet.protection.sheet = False
                    
                    print(f"🔓 Sheet '{sheet_name}' unlocked using Method 2")
                    unlocked_sheets.append(sheet_name)
                    
                except Exception as e2:
                    print(f"⚠️  Method 2 failed for '{sheet_name}': {e2}")
                    
                    # Method 3: Last resort - minimal intervention
                    try:
                        # Just try to disable the main protection flag
                        if hasattr(worksheet, 'protection'):
                            if hasattr(worksheet.protection, 'sheet'):
                                worksheet.protection.sheet = False
                            if hasattr(worksheet.protection, 'password'):
                                worksheet.protection.password = None
                        
                        print(f"🔓 Sheet '{sheet_name}' processed using Method 3")
                        unlocked_sheets.append(sheet_name)
                        
                    except Exception as e3:
                        print(f"❌ All methods failed for sheet '{sheet_name}': {e3}")
                        print("    This sheet will remain as-is in the output file.")
        
        # Report results
        if unlocked_sheets:
            print(f"\n🎉 Successfully unlocked {len(unlocked_sheets)} out of {total_sheets} sheets:")
            for sheet in unlocked_sheets:
                print(f"   - {sheet}")
        else:
            print(f"\n📋 All {total_sheets} sheets were already unlocked")
        
        # Determine output path with cross-platform handling
        if output_path is None:
            file_path_obj = Path(file_path)
            output_path = file_path_obj.parent / f"{file_path_obj.stem}_unlocked{file_path_obj.suffix}"
        else:
            output_path = normalize_path(output_path)
        
        # Ensure output directory exists
        output_dir = Path(output_path).parent
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Save the unlocked workbook
        print(f"\n💾 Saving unlocked file to: {output_path}")
        workbook.save(output_path)
        workbook.close()
        
        print("✅ Process completed successfully!")
        print(f"📁 Unlocked file saved as: {os.path.basename(output_path)}")
        return True
        
    except PermissionError:
        print("❌ Error: Permission denied. Please close Excel and try again.")
        return False
    except FileNotFoundError:
        print("❌ Error: File not found. Please check the file path.")
        return False
    except Exception as e:
        print(f"❌ Error occurred: {str(e)}")
        return False

def process_multiple_files(directory_path):
    """
    Process all Excel files in a directory - Cross platform
    """
    directory_path = normalize_path(directory_path)
    directory = Path(directory_path)
    
    if not directory.exists():
        print(f"❌ Directory not found: {directory_path}")
        return
    
    # Find Excel files with case-insensitive extensions (Windows compatibility)
    excel_files = []
    extensions = ['*.xlsx', '*.xlsm', '*.xls', '*.XLSX', '*.XLSM', '*.XLS']
    
    for pattern in extensions:
        excel_files.extend(directory.glob(pattern))
    
    # Remove duplicates (in case of case-insensitive filesystem)
    excel_files = list(set(excel_files))
    
    if not excel_files:
        print("❌ No Excel files found in the directory")
        return
    
    print(f"📂 Found {len(excel_files)} Excel file(s) to process\n")
    
    successful = 0
    for i, file_path in enumerate(excel_files, 1):
        print(f"--- Processing file {i}/{len(excel_files)}: {file_path.name} ---")
        if unlock_excel_sheets(str(file_path)):
            successful += 1
        print()  # Add spacing between files
    
    print(f"🏁 Final Results: {successful}/{len(excel_files)} files processed successfully")

def get_input_with_prompt(prompt_text, is_path=False):
    """Get user input with cross-platform path handling"""
    user_input = input(prompt_text).strip()
    
    if is_path and user_input:
        return normalize_path(user_input)
    return user_input

def main():
    """
    Main function with cross-platform interface
    """
    sys_info = get_system_info()
    
    print("🔐 Excel Sheet Unlocker - Cross Platform")
    print("=" * 45)
    print(f"🖥️  Running on: {platform.system()} {platform.release()}")
    print("📋 This tool removes sheet protection without needing passwords\n")
    
    # Check for command line arguments
    if len(sys.argv) >= 2:
        input_path = normalize_path(sys.argv[1])
        output_path = normalize_path(sys.argv[2]) if len(sys.argv) >= 3 else None
        
        if os.path.isdir(input_path):
            process_multiple_files(input_path)
        else:
            unlock_excel_sheets(input_path, output_path)
    else:
        # Interactive mode
        print("Choose an option:")
        print("1. Process a single file")
        print("2. Process all Excel files in a folder")
        
        if sys_info['is_windows']:
            print("\n💡 Tip (Windows): You can drag & drop files/folders into this window!")
        elif sys_info['is_macos']:
            print("\n💡 Tip (macOS): You can drag & drop files/folders from Finder!")
        else:
            print("\n💡 Tip (Linux): Copy file paths from your file manager!")
        
        choice = input("\nEnter your choice (1 or 2): ").strip()
        
        if choice == "1":
            if sys_info['is_windows']:
                file_path = get_input_with_prompt("\n📂 Drag & drop Excel file or enter path: ", is_path=True)
            else:
                file_path = get_input_with_prompt("\n📂 Drag & drop Excel file or enter path: ", is_path=True)
            
            custom_output = get_input_with_prompt("💾 Custom output name? (Press Enter to auto-name): ", is_path=True)
            output_path = custom_output if custom_output else None
            
            print()  # Add spacing
            unlock_excel_sheets(file_path, output_path)
            
        elif choice == "2":
            folder_path = get_input_with_prompt("\n📁 Enter folder path containing Excel files: ", is_path=True)
            
            print()  # Add spacing
            process_multiple_files(folder_path)
            
        else:
            print("❌ Invalid choice. Please run the script again.")
    
    # Cross-platform pause before exit
    if sys_info['is_windows']:
        input("\nPress Enter to exit...")

if __name__ == "__main__":
    # Install dependencies with cross-platform support
    if not install_dependencies():
        sys.exit(1)
    
    # Import after installation
    import openpyxl
    
    main()